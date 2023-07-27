import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
import openpyxl


file_path = ""  # Global variable to store the file path
current_sheet = ""  # Global variable to store the selected sheet name
widget_name_mapping = {}  # Global variable to store the mapping of widgets
column_to_widget_mapping = {} 
def load_excel_file():
    global file_path, current_sheet
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx *.xlsm")])
    if file_path:
        current_sheet = ""  # Clear the current sheet when loading a new file
        load_data(file_path)

def load_data(file_path):
    global current_sheet, widget_name_mapping
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet_names = workbook.sheetnames
        
        if not current_sheet or current_sheet not in sheet_names:
            # If current_sheet is empty or not in sheet_names, set it to the first sheet
            current_sheet = sheet_names[0]
            
        # Update the Combobox with sheet names
        sheet_dropdown["values"] = sheet_names
        sheet_dropdown.set(current_sheet)  # Set the selected sheet in the Combobox

        sheet = workbook[current_sheet]
        list_values = list(sheet.iter_rows(values_only=True))
        columns = list_values[0]  # Extract the column headings from the first row of the selected sheet

        # Update the TreeView's headings with the column headings
        update_treeview_headings(columns)

        widget_name_mapping = {}
        column_to_widget_mapping = {}  # Reverse mapping: Column headings to widget names

        for idx, col_name in enumerate(columns):
            if idx < 6:
                widget_name = "widget_" + str(idx)
                widget_name_mapping[widget_name] = col_name
                column_to_widget_mapping[col_name] = widget_name.lower()  # Create the reverse mapping with lowercase
                # Update the label for the corresponding widget
                widgets_frame.grid_slaves(row=0, column=idx)[0].config(text=col_name)

        # Insert the data into the TreeView
        for value_tuple in list_values[1:]:
            treeView.insert('', tk.END, values=value_tuple)

        return sheet_names, column_to_widget_mapping  # Return both sheet names list and the reverse mapping
    
    except openpyxl.utils.exceptions.InvalidFileException:
        print("Invalid Excel file. Please select a valid Excel file.")

        
def on_sheet_select(event):
    global current_sheet
    selected_sheet = sheet_dropdown.get()
    if selected_sheet != current_sheet:
        current_sheet = selected_sheet
        load_data(file_path)  # Update the TreeView when the selected sheet changes

def update_treeview_headings(columns):
    # Clear previous TreeView columns and headings
    treeView.delete(*treeView.get_children())
    treeView["columns"] = columns

    for col_name in columns:
        treeView.heading(col_name, text=col_name)
        treeView.column(col_name, width=100)  # Set a default width for each column (you can adjust it as needed)


def insert_row():
    if not column_to_widget_mapping:
        print("Error: No mapping available.")
    # Retrieve data from the widgets
    checkDate = check_date.get()
    checkNumber = check_number.get()
    dvNumber = dv_number.get()
    checkParticulars = check_particulars.get()
    checkAmount = check_amount.get()
    checkStatus = check_status_dropdown.get()

    # Get the selected sheet from the workbook
    path = file_path
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[current_sheet]
    
    # Insert the row to the selected sheet in the workbook
    row_values = [checkDate, checkNumber, dvNumber, checkParticulars, checkAmount, checkStatus]
    sheet.append(row_values)
    workbook.save(path)

    # displaying the inserted row on the UI (treeView)
    treeView.insert('', tk.END, values=row_values)
    
    # Get the ID of the last inserted row in the TreeView
    last_row_id = treeView.get_children()[-1]

    # clear the values after inserting the new row then resetting the values to default
    check_date.delete(0, "end")
    check_number.delete(0, "end")
    dv_number.delete(0, "end")
    check_particulars.delete(0, "end")
    check_amount.delete(0, "end")
    check_status_dropdown.set(status_list[0])
    
    # Highlight or select the last inserted row in the TreeView
    treeView.selection_set(last_row_id)
    # Scroll the TreeView to make sure the last inserted row is visible
    treeView.see(last_row_id)
    
    # returns the focus to the check_date widget after inserting the new row
    check_date.focus_set()
    
    # Change the style of the "Add" button when it is selected using Tab
    style.map("Custom.TButton",
              foreground=[("active", "white"), ("!active", "black")],
              background=[("active", "blue"), ("!active", "SystemButtonFace")])


# Create the main window
root = tk.Tk()
root.title('MTO Check Transmittal')

style = ttk.Style(root)
root.tk.call("source", "forest-light.tcl")
root.tk.call("source", "forest-dark.tcl")
style.theme_use("forest-dark")

# Create a custom style for the "Add" button
style.configure("Custom.TButton", font=("Arial", 12))

# Update labels of widgets based on column headings and get the widget_name_mapping
sheet_names = load_data(file_path)  # Get the sheet names list

################## Main Frame ##################
'''
    outer_frame.pack() makes the app responsive.
    Since this is the main widget, adjusting the size of the UI/app will keep
    the components centered
'''
outer_frame = ttk.Frame(root) # parent widget
outer_frame.pack()


################## Widgets Frame ##################
''' 
    The user-input form area.
    All other ttk widgets should have "widgets_frame" as their parent
    since they are supposed to be "grouped-together"... except from, of course, the
    widgets_frame for its root will be the outer_frame (see above).
'''

### col0, row0 of the root_frame : Enclosure Widget for all other input widgets
widgets_frame = ttk.LabelFrame(outer_frame, text="Widgets Frame")
widgets_frame.grid(row=0, column=0, padx=20, pady=10) # padding on x & y axis

w00 = ttk.Label(widgets_frame, text="Check Date")
w00.grid(row=0, column=0)

w01 = ttk.Label(widgets_frame, text="Check #")
w01.grid(row=0, column=1)

w02 = ttk.Label(widgets_frame, text="DV #")
w02.grid(row=0, column=2)

w03 = ttk.Label(widgets_frame, text="Particulars")
w03.grid(row=0, column=3)

w04 = ttk.Label(widgets_frame, text="Amount")
w04.grid(row=0, column=4)

w05 = ttk.Label(widgets_frame, text="Status")
w05.grid(row=0, column=5)


check_date = ttk.Entry(widgets_frame, width=10)
# check_date.insert(0, "help text here")
# check_date.bind("<FocusIn>", lambda e: check_date.delete('0', 'end'))
check_date.grid(row=1, column=0,padx=5, pady=(0,5), sticky="ew")

check_number = ttk.Entry(widgets_frame,width=10)
check_number.grid(row=1, column=1,padx=5, pady=(0,5), sticky="ew")

dv_number = ttk.Entry(widgets_frame, width=10)
dv_number.grid(row=1, column=2,padx=5, pady=(0,5), sticky="ew")

check_particulars = ttk.Entry(widgets_frame, width=50)
check_particulars.grid(row=1, column=3,padx=5, pady=(0,5), sticky="ew")

check_amount = ttk.Entry(widgets_frame, width=20)
check_amount.grid(row=1, column=4,padx=5, pady=(0,5), sticky="ew")

status_list = ["Paid", "Cancelled", "Stale"]
check_status_dropdown = ttk.Combobox(widgets_frame, values=status_list)
check_status_dropdown.current(0) # default selected value
check_status_dropdown.grid(row=1, column=5,padx=5, pady=(0,5), sticky="ew")

# Select File Button
btn_load = ttk.Button(widgets_frame, text="Load Excel File", command=load_excel_file, takefocus=0)
btn_load.grid(row=0, column=6, padx=5, pady=(0, 5), sticky="ew")

# Select Sheet Button (Inside Widgets Frame)
def select_sheet():
    selected_item = treeView.focus()
    if selected_item:
        item_values = treeView.item(selected_item, "values")
        selected_sheet_name = item_values[0]  # Assuming the first value is the "Check Date" field
        print(f"Selected sheet: {selected_sheet_name}")

# Dropdown menu for selecting sheets
sheet_dropdown = ttk.Combobox(widgets_frame, state="readonly")
sheet_dropdown.grid(row=4, column=5, padx=10, pady=5)
sheet_dropdown.bind("<<ComboboxSelected>>", on_sheet_select)

# Insert Row button (Inside Widgets Frame)
btn_row = tk.Button(widgets_frame, text="Boss Phat_05 sakalam!", command=lambda: insert_row(), takefocus=1)
btn_row.grid(row=4, column=3, sticky="nsew")


################## TreeView / Excel LabelFrame ####################################
### This is where the preview of the excel file's data will be displayed

### Display Frame
treeFrame = ttk.Frame(outer_frame, takefocus=0)
treeFrame.grid(row=5, column=0, pady=10)
treeScroll = ttk.Scrollbar(treeFrame)
treeScroll.pack(side="right", fill="y") # this sets the scrollbar to the right side of the frame,
                                        # covering its whole height
cols = ("Check Date", "Check #", "DV #","Particulars", "Amount", "Status") # column of the preview related to the excel file
treeView = ttk.Treeview(treeFrame, show="headings", 
                        yscrollcommand=treeScroll.set, columns=cols, height=15)
# these set the width of the columns specifically
# COLUMN NAMES SHOULD MATCH THOSE INDICATED ON THE EXCEL FILE
treeView.column("Check Date", width=100)
treeView.column("Check #", width=100)
treeView.column("DV #", width=100)
treeView.column("Particulars", width=200)
treeView.column("Amount", width=100)
treeView.column("Status", width=80)
treeView.pack()
treeScroll.config(command=treeView.yview) # this line attaches the treeScroll widget to the treeView, scrolling vertically


# Event Listener function highlighting selected items on the treeView list
def selected():
    print(listbox.get(listbox.curselection()[0]))
    
treeView.bind("<<ListboxSelect>>", lambda x: selected())

### switch (dark/light)
def toggle_mode():
    if mode_switch.instate(["selected"]):
        style.theme_use("forest-light")
    else:
        style.theme_use("forest-dark")

mode_switch = ttk.Checkbutton(outer_frame, text="Mode", style="Switch",
    command=toggle_mode, takefocus=0) # this triggers the toggle_mode function above
mode_switch.grid(row=6, column=0, padx=5, pady=10, sticky="nsew")

check_date.focus_set()

root.mainloop()
"""_summary_
"""